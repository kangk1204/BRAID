"""Tests for the caller-agnostic BRAID filter layer (braid/adapters + braid filter).

In-memory only: synthetic native-format snippets for each caller, a stub
calibrator for the tier logic, and the packaged calibrator for the end-to-end
path (skipped if the artifact is absent).
"""
from __future__ import annotations

import argparse
import os
from pathlib import Path

import pytest

from braid.adapters import (
    CallerEvent,
    beta_dpsi_std,
    calibrate_event,
    calibrate_events,
    parse_betas,
    parse_majiq,
    parse_rmats,
    parse_suppa2,
)

_RMATS_HEADER = [
    "ID", "GeneID", "geneSymbol", "chr", "strand", "exonStart_0base", "exonEnd",
    "upstreamES", "upstreamEE", "downstreamES", "downstreamEE",
    "IJC_SAMPLE_1", "SJC_SAMPLE_1", "IJC_SAMPLE_2", "SJC_SAMPLE_2",
    "IncFormLen", "SkipFormLen", "PValue", "FDR",
    "IncLevel1", "IncLevel2", "IncLevelDifference",
]


# --------------------------------------------------------------------------- #
# Stub calibrator: a fixed absolute half-width, so tier logic is deterministic.
# --------------------------------------------------------------------------- #
class _StubCal:
    def __init__(self, half: float = 0.3) -> None:
        self.half = half

    def interval(self, dpsi, sigma, support, *, event_type=None,
                 force_global=False, clip=(-1.0, 1.0)):
        return max(dpsi - self.half, clip[0]), min(dpsi + self.half, clip[1])

    def robust_interval(self, dpsi, std, support, *, event_type=None,
                        force_global=False, clip=(-1.0, 1.0)):
        h = (self.half ** 2 + (1.959963984540054 * std) ** 2) ** 0.5
        return max(dpsi - h, clip[0]), min(dpsi + h, clip[1])

    def check_applicability(self, supports, **kw):
        return True, ""


def _ev(dpsi, *, fdr=None, pvalue=None, std=None, et="SE", support=300.0):
    return CallerEvent(
        event_id="e", gene="G", event_type=et, dpsi=dpsi,
        total_support=support, caller="test", fdr=fdr, pvalue=pvalue,
        sampling_std=std,
    )


# --------------------------------------------------------------------------- #
# base
# --------------------------------------------------------------------------- #
def test_beta_dpsi_std_decreases_with_depth():
    shallow = beta_dpsi_std(9, 1, 1, 9)
    deep = beta_dpsi_std(900, 100, 100, 900)
    assert deep < shallow
    assert shallow > 0.0


@pytest.mark.parametrize(
    "dpsi, fdr, pvalue, expect",
    [
        (0.6, 0.001, None, "high-confidence"),     # reliable + effect + caller-sig
        (0.6, 0.80, None, "supported"),            # reliable + effect, caller not sig
        (0.05, 0.001, None, "caller-significant-only"),  # caller-sig but interval crosses 0
        (0.05, 0.80, None, "not-significant"),     # neither
        (0.6, None, None, "high-confidence"),      # caller gave no sig -> BRAID-only
        (0.05, None, None, "not-reliable"),        # caller gave no sig, interval crosses 0
    ],
)
def test_calibrate_event_tiers(dpsi, fdr, pvalue, expect):
    r = calibrate_event(_ev(dpsi, fdr=fdr, pvalue=pvalue), _StubCal(0.3))
    assert r["tier"] == expect


def test_calibrate_events_sorted_reliable_first():
    evs = [_ev(0.05, fdr=0.8), _ev(0.6, fdr=0.001), _ev(-0.5, fdr=0.001)]
    rows = calibrate_events(evs, _StubCal(0.3))
    # both reliable events (|dpsi|=0.6, 0.5) sort above the zero-crossing one
    assert rows[0]["reliable"] and rows[1]["reliable"]
    assert not rows[-1]["reliable"]


def test_robust_interval_used_when_sampling_std_present():
    narrow = calibrate_event(_ev(0.5, std=0.0), _StubCal(0.3))
    wide = calibrate_event(_ev(0.5, std=0.2), _StubCal(0.3))
    assert (wide["ci_high"] - wide["ci_low"]) > (narrow["ci_high"] - narrow["ci_low"])


# --------------------------------------------------------------------------- #
# parsers
# --------------------------------------------------------------------------- #
def test_parse_majiq(tmp_path: Path):
    p = tmp_path / "deltapsi.tsv"
    p.write_text(
        "gene_name\tlsv_id\tE(dPSI) per LSV junction\tP(|dPSI|>=0.20) per LSV junction\tnum_reads\n"
        "TRA2B\tg1:s\t0.10;0.45\t0.30;0.97\t220\n"
        "QKI\tg2:t\t-0.05\t0.40\t180\n"
    )
    evs = parse_majiq(str(p))
    assert len(evs) == 2
    # most-changing junction summary picks |0.45| over |0.10|
    assert evs[0].dpsi == pytest.approx(0.45)
    assert evs[0].event_type == "LSV"
    # P>=0.97 maps to pvalue 1-0.97 = 0.03 (significant)
    assert evs[0].pvalue == pytest.approx(0.03, abs=1e-6)


def test_parse_suppa2_unnamed_id_column(tmp_path: Path):
    p = tmp_path / "diff.dpsi"
    # SUPPA2 leaves the id column header blank -> data rows have one extra field.
    p.write_text(
        "Ctrl-KD_dPSI\tCtrl-KD_p-val\n"
        "ENSG01;SE:chr1:100-200:+\t0.40\t0.002\n"
        "ENSG02;A3:chr2:50-90:-\t-0.30\t0.500\n"
    )
    evs = parse_suppa2(str(p))
    assert len(evs) == 2
    assert evs[0].dpsi == pytest.approx(0.40)
    assert evs[0].pvalue == pytest.approx(0.002)
    assert evs[0].event_type == "SE"
    assert evs[1].event_type == "A3SS"          # A3 -> A3SS mapping
    assert evs[0].total_support is None          # SUPPA2 reports no read support


def test_parse_betas(tmp_path: Path):
    p = tmp_path / "betas.tsv"
    p.write_text(
        "gene\tevent_id\tdeltapsi\tlower\tupper\tfdr\n"
        "BIG\tBIG:1\t0.55\t0.40\t0.70\t0.01\n"
        "NULL\tNULL:1\t0.02\t-0.20\t0.24\t0.90\n"
    )
    evs = parse_betas(str(p))
    assert len(evs) == 2
    assert evs[0].dpsi == pytest.approx(0.55)
    assert evs[0].caller_low == pytest.approx(0.40)
    # sampling_std derived from interval width (0.70-0.40)/(2*1.96)
    assert evs[0].sampling_std == pytest.approx(0.30 / (2 * 1.959963984540054), abs=1e-6)


def test_parse_betas_event_id_does_not_collide_with_event_type(tmp_path: Path):
    """A betAS table with event_type but no event_id must not give every event the
    event_type as its id (identifier corruption that collapses downstream joins)."""
    p = tmp_path / "b.tsv"
    p.write_text(
        "gene\tevent_type\tdeltapsi\tlower\tupper\n"
        "MYG\tA3SS\t0.5\t0.4\t0.6\n"
        "OTH\tA3SS\t-0.4\t-0.5\t-0.3\n"
    )
    evs = parse_betas(str(p))
    assert [e.event_type for e in evs] == ["A3SS", "A3SS"]   # type still parsed
    assert [e.event_id for e in evs] != ["A3SS", "A3SS"]     # id is NOT the type
    assert evs[0].event_id == "MYG" and evs[1].event_id == "OTH"  # falls back to gene


def test_parse_betas_missing_dpsi_raises(tmp_path: Path):
    p = tmp_path / "bad.tsv"
    p.write_text("gene\tlower\tupper\nX\t0.1\t0.2\n")
    with pytest.raises(ValueError, match="no ΔPSI column"):
        parse_betas(str(p))


def _write_rmats(rmats_dir: Path) -> None:
    rmats_dir.mkdir(parents=True, exist_ok=True)

    def row(eid, gene, inc1, sjc1, inc2, sjc2, fdr, dpsi):
        psi1, psi2 = inc1 / (inc1 + sjc1), inc2 / (inc2 + sjc2)
        return [
            str(eid), gene, gene, "chr1", "+", "1000", "1100",
            "800", "900", "1200", "1300",
            str(inc1), str(sjc1), str(inc2), str(sjc2),
            "100", "100", "0.0", str(fdr),
            f"{psi1:.3f}", f"{psi2:.3f}", str(dpsi),
        ]
    rows = [
        row(1, "BIG", 180, 20, 60, 140, 0.001, 0.6),
        row(2, "NULL", 100, 100, 100, 100, 0.8, 0.0),
    ]
    lines = ["\t".join(_RMATS_HEADER)] + ["\t".join(r) for r in rows]
    (rmats_dir / "SE.MATS.JC.txt").write_text("\n".join(lines) + "\n")


def test_parse_rmats(tmp_path: Path):
    _write_rmats(tmp_path / "rmats")
    evs = parse_rmats(str(tmp_path / "rmats"), min_support=20)
    assert len(evs) == 2
    big = next(e for e in evs if e.gene == "BIG")
    assert big.dpsi == pytest.approx(0.6)
    assert big.total_support == pytest.approx(400.0)   # 180+20+60+140
    assert big.sampling_std is not None and big.sampling_std > 0.0
    assert big.event_type == "SE"


def test_parse_rmats_drops_na_inclevel_difference(tmp_path: Path, caplog) -> None:
    """rMATS IncLevelDifference=NA must be dropped with a warning, not emitted as a
    NaN ΔPSI -- which downstream becomes a full [-1, 1] interval and a literal 'nan'
    in the output TSV. Matches the MAJIQ/SUPPA2/betAS out-of-domain handling."""
    import math

    rmats_dir = tmp_path / "rmats"
    rmats_dir.mkdir(parents=True, exist_ok=True)

    def row(eid: int, gene: str, dpsi: object) -> list[str]:
        # Valid counts in both groups; only IncLevelDifference varies.
        return [
            str(eid), gene, gene, "chr1", "+", "1000", "1100",
            "800", "900", "1200", "1300",
            "50", "50", "30", "70", "100", "100", "0.5", "0.05",
            "0.500", "0.300", str(dpsi),
        ]

    lines = ["\t".join(_RMATS_HEADER)] + [
        "\t".join(row(1, "OK", 0.2)),
        "\t".join(row(2, "BAD", "NA")),
    ]
    (rmats_dir / "SE.MATS.JC.txt").write_text("\n".join(lines) + "\n")

    with caplog.at_level("WARNING", logger="braid.adapters.parsers"):
        evs = parse_rmats(str(rmats_dir), min_support=20)

    assert [e.gene for e in evs] == ["OK"]           # the NA row is dropped
    assert all(math.isfinite(e.dpsi) for e in evs)   # no NaN ΔPSI emitted
    assert any("dropped" in r.message for r in caplog.records)


# --------------------------------------------------------------------------- #
# end-to-end with the packaged calibrator + report writers
# --------------------------------------------------------------------------- #
@pytest.fixture
def real_calibrator():
    from braid.target.conformal import load_differential_conformal_calibrator
    try:
        return load_differential_conformal_calibrator()
    except (FileNotFoundError, OSError):
        pytest.skip("packaged differential calibrator artifact not available")


def test_end_to_end_rmats_real_calibrator(tmp_path: Path, real_calibrator):
    _write_rmats(tmp_path / "rmats")
    evs = parse_rmats(str(tmp_path / "rmats"), min_support=20)
    rows = calibrate_events(evs, real_calibrator)
    big = next(r for r in rows if r["gene"] == "BIG")
    assert -1.0 <= big["ci_low"] <= big["ci_high"] <= 1.0
    assert big["tier"] in {
        "high-confidence", "supported", "caller-significant-only",
        "not-significant", "not-reliable",
    }


def test_report_writers(tmp_path: Path, real_calibrator):
    from braid.adapters.report import make_figure, write_excel, write_tsv
    _write_rmats(tmp_path / "rmats")
    rows = calibrate_events(
        parse_rmats(str(tmp_path / "rmats"), min_support=20), real_calibrator
    )
    write_tsv(rows, str(tmp_path / "out.tsv"))
    assert (tmp_path / "out.tsv").read_text().startswith("tier\t")
    if write_excel(rows, str(tmp_path / "out.xlsx"), caller="rmats"):
        assert (tmp_path / "out.xlsx").exists()
    if make_figure(rows, str(tmp_path / "fig"), caller="rmats"):
        assert (tmp_path / "fig.png").exists()


def test_filter_output_neutralises_formula_injection(tmp_path: Path):
    """Gene/event ids starting with a formula trigger must be rendered inert in the
    TSV and XLSX (spreadsheet formula-injection hardening)."""
    from braid.adapters.report import write_excel, write_tsv
    rows = [{
        "tier": "high-confidence", "gene": '=HYPERLINK("http://evil","x")',
        "event_id": "@SUM(1+1)", "event_type": "SE", "caller": "majiq",
        "dpsi": 0.5, "ci_low": 0.1, "ci_high": 0.9, "reliable": True,
        "caller_significant": True, "pvalue": None, "fdr": 0.01,
        "total_support": 100.0, "support_known": True,
        "group1_psi": None, "group2_psi": None, "caller_low": None, "caller_high": None,
    }]
    tsv = tmp_path / "o.tsv"
    write_tsv(rows, str(tsv))
    body = tsv.read_text()
    assert "\t'=HYPERLINK" in body and "\t'@SUM" in body   # neutralised
    assert "\t=HYPERLINK" not in body                       # no raw formula cell

    xlsx = tmp_path / "o.xlsx"
    if write_excel(rows, str(xlsx), caller="majiq"):
        from openpyxl import load_workbook
        ws = load_workbook(xlsx)["calibrated"]
        starts = [c.value for c in ws[2] if isinstance(c.value, str)]
        assert not any(s.startswith(("=", "+", "@")) for s in starts)


def test_cli_filter_suppa2(tmp_path: Path, real_calibrator):
    from braid.commands.filter_cmd import run_filter
    dpsi = tmp_path / "diff.dpsi"
    dpsi.write_text(
        "Ctrl-KD_dPSI\tCtrl-KD_p-val\n"
        "ENSG01;SE:chr1:100-200:+\t0.40\t0.002\n"
        "ENSG02;SE:chr2:50-90:-\t0.02\t0.500\n"
    )
    out = tmp_path / "braid_filter"
    args = argparse.Namespace(
        input=str(dpsi), caller="suppa2", output=str(out),
        effect_cutoff=0.1, sig_threshold=0.05, min_support=20,
        calibration=None, top_n=25, make_figure=False, verbose=False,
    )
    run_filter(args)
    assert (tmp_path / "braid_filter.tsv").exists()
    body = (tmp_path / "braid_filter.tsv").read_text()
    assert "high-confidence" in body or "caller-significant-only" in body


def test_cli_filter_rejects_invalid_top_n(tmp_path: Path):
    """--top-n < 1 must fail fast. It slices the figure short-list as rows[:top_n], so
    a negative value silently drops events (rows[:-1] == all but last) and 0 yields an
    empty figure."""
    from braid.commands.filter_cmd import run_filter
    dpsi = tmp_path / "diff.dpsi"
    dpsi.write_text(
        "Ctrl-KD_dPSI\tCtrl-KD_p-val\n"
        "ENSG01;SE:chr1:100-200:+\t0.40\t0.002\n"
    )
    out = tmp_path / "braid_filter"

    def _args(top_n: int) -> argparse.Namespace:
        return argparse.Namespace(
            input=str(dpsi), caller="suppa2", output=str(out),
            effect_cutoff=0.1, sig_threshold=0.05, min_support=20,
            calibration=None, top_n=top_n, make_figure=False, verbose=False,
        )

    for bad in (0, -1, -25):
        with pytest.raises(ValueError, match="top-n"):
            run_filter(_args(bad))


_EXAMPLES = os.path.join(
    os.path.dirname(os.path.dirname(__file__)), "braid", "examples", "filter"
)


@pytest.mark.parametrize(
    "caller, rel",
    [
        ("rmats", "rmats_output"),
        ("majiq", "majiq_deltapsi.tsv"),
        ("suppa2", "suppa2_diffSplice.dpsi"),
        ("betas", "betas_differential.tsv"),
    ],
)
def test_bundled_examples_give_expected_tiers(caller, rel, real_calibrator):
    """The shipped per-caller samples must keep producing the documented tiers."""
    from braid.adapters import PARSERS, calibrate_events
    path = os.path.join(_EXAMPLES, rel)
    assert os.path.exists(path), f"missing bundled example: {path}"
    parser = PARSERS[caller]
    events = parser(path, min_support=0) if caller == "rmats" else parser(path)
    rows = calibrate_events(events, real_calibrator)
    tier_by = {r["gene"]: r["tier"] for r in rows}

    def tier_of(sub: str) -> str:
        return next(t for g, t in tier_by.items() if sub in g)

    assert tier_of("BIG") == "high-confidence"
    assert tier_of("MOD") == "caller-significant-only"
    assert tier_of("NULL") == "not-significant"


def test_parse_majiq_pairs_dpsi_and_prob_from_same_junction(tmp_path: Path):
    """The reported dPSI and probability must come from the SAME junction, not be
    selected independently (a large effect on one junction mixed with a high
    probability on another would falsely inflate the tier)."""
    p = tmp_path / "m.tsv"
    p.write_text(
        "gene_name\tlsv_id\tE(dPSI) per LSV junction\tP(|dPSI|>=0.20) per LSV junction\n"
        "G\tg:1\t0.80;0.10\t0.10;0.99\n"
    )
    ev = parse_majiq(str(p))[0]
    assert ev.dpsi == pytest.approx(0.80)            # most-changing junction
    assert ev.pvalue == pytest.approx(1.0 - 0.10)    # ITS probability (0.10), not 0.99


def test_parse_suppa2_warns_on_multiple_contrasts(tmp_path: Path, caplog):
    import logging
    p = tmp_path / "diff.dpsi"
    p.write_text(
        "A-B_dPSI\tA-B_p-val\tC-D_dPSI\tC-D_p-val\n"
        "ENSG;SE:chr1:1-2:+\t0.3\t0.01\t0.9\t0.5\n"
    )
    with caplog.at_level(logging.WARNING):
        evs = parse_suppa2(str(p))
    assert evs[0].dpsi == pytest.approx(0.3)         # first contrast used
    assert any("dPSI contrast" in r.message for r in caplog.records)


def test_filter_rejects_wrong_scale_kind_calibrator(tmp_path: Path):
    """A custom calibrator fit for the wrong mode must be rejected, not silently
    misinterpreted (posterior_std vs absolute_dpsi change the interval units)."""
    from braid.commands.filter_cmd import _resolve_calibrator
    from braid.target.conformal import ConformalCalibrator
    wrong = ConformalCalibrator(
        alpha=0.05, q_global=0.3, q_by_bin={}, scale_kind="posterior_std")
    wp = tmp_path / "wrong.json"
    wrong.to_json(wp)
    with pytest.raises(ValueError, match="scale_kind"):
        _resolve_calibrator(str(wp))
    right = ConformalCalibrator(
        alpha=0.05, q_global=0.3, q_by_bin={}, scale_kind="absolute_dpsi")
    rp = tmp_path / "right.json"
    right.to_json(rp)
    assert _resolve_calibrator(str(rp)).scale_kind == "absolute_dpsi"


def _bin_aware_cal():
    """A custom calibrator that defines every cascade level, so a misrouted event
    would visibly pick a tighter quantile than the global one."""
    from braid.target.conformal import ConformalCalibrator
    return ConformalCalibrator(
        alpha=0.05, q_global=0.5,
        q_by_bin={"<20": 0.01, "250+": 0.05},
        q_by_event_type={"SE": 0.02},
        q_by_group={"SE|<20": 0.03},
        scale_kind="absolute_dpsi",
    )


def test_unknown_support_uses_global_not_tightest_bin():
    """SUPPA2/MAJIQ (no read support, total_support=None) events must use the pooled
    global quantile, never a support/event-type/composite bin -- even on a custom
    calibrator that defines all of them (an explicit contract, not a nan side-effect)."""
    cal = _bin_aware_cal()
    ev = CallerEvent(event_id="e", gene="G", event_type="SE", dpsi=0.3,
                     total_support=None, caller="suppa2")
    r = calibrate_event(ev, cal)
    half = (r["ci_high"] - r["ci_low"]) / 2
    assert r["support_known"] is False
    assert r["total_support"] is None
    assert half == pytest.approx(0.5, abs=1e-6)   # global q, not 0.01/0.02/0.03/0.05


def test_real_support_of_1000_is_known_and_uses_support_bin():
    """Regression: a count-bearing caller whose real total support is exactly 1000
    must stay support-known and use its support bin -- the old numeric sentinel
    (1000.0) would have silently mislabelled it unknown and forced the global q."""
    cal = _bin_aware_cal()
    ev = CallerEvent(event_id="e", gene="G", event_type="RI", dpsi=0.3,
                     total_support=1000.0, caller="betas")
    r = calibrate_event(ev, cal)
    half = (r["ci_high"] - r["ci_low"]) / 2
    assert r["support_known"] is True
    assert r["total_support"] == pytest.approx(1000.0)
    # support 1000 -> "250+" bin; event_type RI has no q -> bin q 0.05 (NOT global 0.5)
    assert half == pytest.approx(0.05, abs=1e-6)
