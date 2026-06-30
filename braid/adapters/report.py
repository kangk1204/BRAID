"""Output writers for the caller-agnostic filter: TSV, Excel, and a figure.

The TSV is always written (no extra dependency). The Excel workbook and the
publication-ready figure need the ``report`` extra (``pip install braid[report]``)
and degrade to a friendly note if their libraries are absent, so the core path
never fails for lack of an optional dependency.
"""

from __future__ import annotations

import logging

from braid.output_safety import csv_safe

logger = logging.getLogger(__name__)

COLUMNS = [
    "tier", "gene", "event_id", "event_type", "caller",
    "dpsi", "ci_low", "ci_high", "reliable",
    "caller_significant", "pvalue", "fdr",
    "total_support", "support_known",
    "group1_psi", "group2_psi", "caller_low", "caller_high",
]

TIER_COLORS = {
    "high-confidence": "#1b6ca8",
    "supported": "#4a9b6e",
    "caller-significant-only": "#d1a32b",
    "not-significant": "#9aa0a6",
    "not-reliable": "#9aa0a6",
}
TIER_ORDER = [
    "high-confidence", "supported", "caller-significant-only",
    "not-significant", "not-reliable",
]


def _fmt(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, float):
        return f"{value:.6f}"
    return csv_safe(str(value))


def write_tsv(rows: list[dict], path: str) -> None:
    """Write the calibrated events to a TSV (always; header even when empty)."""
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\t".join(COLUMNS) + "\n")
        for r in rows:
            fh.write("\t".join(_fmt(r.get(c)) for c in COLUMNS) + "\n")


def tier_counts(rows: list[dict]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for r in rows:
        counts[r["tier"]] = counts.get(r["tier"], 0) + 1
    return counts


def write_excel(rows: list[dict], path: str, *, caller: str = "") -> bool:
    """Write a multi-sheet workbook (calibrated / reliable / summary).

    Returns ``True`` if written, ``False`` if openpyxl is unavailable.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        logger.warning(
            "openpyxl not installed; skipping Excel output "
            "(`pip install braid[report]`). TSV was still written."
        )
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "calibrated"
    _fill_sheet(ws, rows, Font, PatternFill)

    reliable = [r for r in rows if r["reliable"]]
    _fill_sheet(wb.create_sheet("reliable"), reliable, Font, PatternFill)

    summary = wb.create_sheet("summary")
    summary["A1"] = "BRAID calibrated filter — summary"
    summary["A1"].font = Font(bold=True, size=13)
    summary["A3"] = "caller"
    summary["B3"] = caller
    summary["A4"] = "events"
    summary["B4"] = len(rows)
    summary["A5"] = "reliable (interval excludes 0)"
    summary["B5"] = len(reliable)
    summary["A7"] = "tier"
    summary["B7"] = "count"
    summary["A7"].font = summary["B7"].font = Font(bold=True)
    counts = tier_counts(rows)
    for i, tier in enumerate([t for t in TIER_ORDER if t in counts], start=8):
        summary.cell(row=i, column=1, value=tier)
        summary.cell(row=i, column=2, value=counts[tier])
    wb.save(path)
    return True


def _fill_sheet(ws, rows: list[dict], Font, PatternFill) -> None:
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([_excel_value(r.get(c)) for c in COLUMNS])
    # Colour the tier cell (column 1) by tier for at-a-glance reading.
    for i, r in enumerate(rows, start=2):
        hexc = TIER_COLORS.get(r["tier"], "#ffffff").lstrip("#")
        ws.cell(row=i, column=1).fill = PatternFill(
            start_color=hexc, end_color=hexc, fill_type="solid"
        )
    for col, width in (("A", 22), ("B", 16), ("C", 28), ("D", 11)):
        ws.column_dimensions[col].width = width


def _excel_value(value):
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, str):
        return csv_safe(value)
    return value


def make_figure(rows: list[dict], path_stem: str, *, caller: str = "", top_n: int = 25) -> bool:
    """Write a two-panel publication figure (png/pdf/svg).

    (A) Calibrated ΔPSI with 95% intervals for the top-N most confidently non-zero
    events, coloured by tier. (B) Tier composition. Returns ``False`` if matplotlib
    is unavailable.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.patches import Patch
    except ImportError:
        logger.warning(
            "matplotlib not installed; skipping figure "
            "(`pip install braid[report]`)."
        )
        return False
    if not rows:
        return False

    plt.rcParams.update(
        {"font.size": 9, "axes.titlesize": 10, "font.family": "DejaVu Sans"}
    )
    fig, (axL, axR) = plt.subplots(
        1, 2, figsize=(11, 5.2), gridspec_kw={"width_ratios": [2.0, 1.0]}
    )

    top = rows[:top_n]
    ys = list(range(len(top)))[::-1]
    for r, y in zip(top, ys):
        color = TIER_COLORS.get(r["tier"], "#9aa0a6")
        axL.plot([r["ci_low"], r["ci_high"]], [y, y], "-", color=color, lw=2.0, alpha=0.85)
        axL.plot([r["dpsi"]], [y], "o", color=color, ms=5)
    axL.axvline(0.0, color="#444", ls="--", lw=1)
    axL.set_yticks(ys)
    axL.set_yticklabels([f"{r['gene'] or r['event_id'][:14]}" for r in top], fontsize=7)
    axL.set_xlabel("ΔPSI (calibrated 95% interval)")
    axL.set_title(f"(A) Top {len(top)} events by calibrated confidence")
    axL.set_xlim(-1.05, 1.05)

    counts = tier_counts(rows)
    tiers = [t for t in TIER_ORDER if t in counts]
    axR.bar(
        range(len(tiers)), [counts[t] for t in tiers],
        color=[TIER_COLORS[t] for t in tiers],
    )
    axR.set_xticks(range(len(tiers)))
    axR.set_xticklabels([t.replace("-", "-\n") for t in tiers], fontsize=7, rotation=0)
    axR.set_ylabel("events")
    axR.set_title("(B) Confidence tiers")
    for i, t in enumerate(tiers):
        axR.text(i, counts[t], str(counts[t]), ha="center", va="bottom", fontsize=8)

    legend = [
        Patch(color=TIER_COLORS[t], label=t) for t in TIER_ORDER
        if t in counts
    ]
    fig.legend(handles=legend, loc="lower center", ncol=len(legend),
               fontsize=7, frameon=False, bbox_to_anchor=(0.5, -0.02))
    fig.suptitle(
        f"BRAID calibrated splicing confidence ({caller}, n={len(rows)})",
        fontsize=11, y=1.0,
    )
    fig.tight_layout(rect=(0, 0.03, 1, 0.98))
    for ext in ("png", "pdf", "svg"):
        fig.savefig(f"{path_stem}.{ext}", dpi=300 if ext == "png" else None,
                    bbox_inches="tight")
    plt.close(fig)
    return True
